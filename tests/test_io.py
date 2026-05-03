from __future__ import annotations

import os
import tempfile
import unittest

from openpyxl import load_workbook

from utils.io import create_poison_rate_directory, export_experiment_results


class ExportResultsTestCase(unittest.TestCase):
    def _build_summary_records(self) -> list[dict[str, object]]:
        return [
            {
                "experiment_name": "ToyExperiment",
                "model": "toy",
                "dataset": "toyset",
                "partition_type": "iid",
                "dirichlet_alpha": None,
                "attack_name": "badnets",
                "attack_mode": "targeted",
                "run_type": "robust",
                "method": "fedavg",
                "feature_mode": "raw_full",
                "mean_f1": float("nan"),
                "mean_bm_gap": float("nan"),
                "final_acc": 0.8123,
                "final_asr": 0.6789,
                "final_bm_gap": float("nan"),
                "malicious_client_ids": "[1, 2]",
            },
            {
                "experiment_name": "ToyExperiment",
                "model": "toy",
                "dataset": "toyset",
                "partition_type": "iid",
                "dirichlet_alpha": None,
                "attack_name": "badnets",
                "attack_mode": "targeted",
                "run_type": "robust",
                "method": "krum",
                "feature_mode": "raw_full",
                "mean_f1": 0.4,
                "mean_bm_gap": 0.2,
                "final_acc": 0.9234,
                "final_asr": 0.1234,
                "final_bm_gap": 0.2,
                "malicious_client_ids": "[1, 2]",
            },
            {
                "experiment_name": "ToyExperiment",
                "model": "toy",
                "dataset": "toyset",
                "partition_type": "iid",
                "dirichlet_alpha": None,
                "attack_name": "badnets",
                "attack_mode": "targeted",
                "run_type": "robust",
                "method": "krum",
                "feature_mode": "selected_layers_balanced",
                "mean_f1": 0.5,
                "mean_bm_gap": 0.25,
                "final_acc": 0.9345,
                "final_asr": 0.2222,
                "final_bm_gap": 0.25,
                "malicious_client_ids": "[1, 2]",
            },
            {
                "experiment_name": "ToyExperiment",
                "model": "toy",
                "dataset": "toyset",
                "partition_type": "iid",
                "dirichlet_alpha": None,
                "attack_name": "badnets",
                "attack_mode": "targeted",
                "run_type": "robust",
                "method": "krum",
                "feature_mode": "selected_layers_balanced_projected",
                "mean_f1": 0.6,
                "mean_bm_gap": 0.3,
                "final_acc": 0.9456,
                "final_asr": 0.2345,
                "final_bm_gap": 0.3,
                "malicious_client_ids": "[1, 2]",
            },
            {
                "experiment_name": "ToyExperiment",
                "model": "toy",
                "dataset": "toyset",
                "partition_type": "iid",
                "dirichlet_alpha": None,
                "attack_name": "label_flipping_untargeted",
                "attack_mode": "untargeted",
                "run_type": "robust",
                "method": "fedavg",
                "feature_mode": "raw_full",
                "mean_f1": float("nan"),
                "mean_bm_gap": float("nan"),
                "final_acc": 0.8123,
                "final_asr": float("nan"),
                "final_bm_gap": float("nan"),
                "malicious_client_ids": "[1, 2]",
            },
            {
                "experiment_name": "ToyExperiment",
                "model": "toy",
                "dataset": "toyset",
                "partition_type": "iid",
                "dirichlet_alpha": None,
                "attack_name": "label_flipping_untargeted",
                "attack_mode": "untargeted",
                "run_type": "robust",
                "method": "krum",
                "feature_mode": "raw_full",
                "mean_f1": 0.7,
                "mean_bm_gap": 0.4,
                "final_acc": 0.4567,
                "final_asr": float("nan"),
                "final_bm_gap": 0.4,
                "malicious_client_ids": "[1, 2]",
            },
            {
                "experiment_name": "ToyExperiment",
                "model": "toy",
                "dataset": "toyset",
                "partition_type": "iid",
                "dirichlet_alpha": None,
                "attack_name": "label_flipping_untargeted",
                "attack_mode": "untargeted",
                "run_type": "robust",
                "method": "krum",
                "feature_mode": "selected_layers_balanced_projected",
                "mean_f1": 0.8,
                "mean_bm_gap": 0.5,
                "final_acc": 0.5678,
                "final_asr": float("nan"),
                "final_bm_gap": 0.5,
                "malicious_client_ids": "[1, 2]",
            },
        ]

    def _export_toy_results(self, output_dir: str, **kwargs: object) -> None:
        export_experiment_results(
            output_dir=output_dir,
            experiment_name="ToyExperiment",
            summary_records=self._build_summary_records(),
            round_logs=[],
            krum_score_logs=[],
            attacks=["badnets", "label_flipping_untargeted"],
            methods=["krum", "fedavg"],
            feature_modes=[
                "raw_full",
                "selected_layers_balanced",
                "selected_layers_balanced_projected",
            ],
            method_display_names={
                "fedavg": "FedAvg",
                "krum": "Krum",
            },
            feature_display_names={
                "raw_full": "原始",
                "selected_layers_balanced": "提取",
                "selected_layers_balanced_projected": "提取+投影",
            },
            export_excel=True,
            **kwargs,
        )

    def test_create_poison_rate_directory_uses_named_subdirectory(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = create_poison_rate_directory(temp_dir, 0.2)

            self.assertTrue(os.path.isdir(output_dir))
            self.assertTrue(output_dir.endswith("poison_rate_0p2"))

    def test_export_experiment_results_uses_report_workbook_layout(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            self._export_toy_results(temp_dir)

            excel_path = os.path.join(temp_dir, "ToyExperiment.xlsx")
            workbook = load_workbook(excel_path, data_only=False)
            self.assertEqual(workbook.sheetnames, ["BMGap", "ACC", "ASR"])

            bm_gap_sheet = workbook["BMGap"]
            self.assertEqual(bm_gap_sheet["B1"].value, "Krum")
            self.assertEqual([bm_gap_sheet.cell(row=2, column=col).value for col in range(2, 5)], ["原始", "提取", "提取+投影"])
            self.assertEqual(bm_gap_sheet["A3"].value, "badnets")
            self.assertAlmostEqual(float(bm_gap_sheet["B3"].value), 0.2, places=4)
            self.assertAlmostEqual(float(bm_gap_sheet["C3"].value), 0.25, places=4)
            self.assertAlmostEqual(float(bm_gap_sheet["D3"].value), 0.3, places=4)
            self.assertEqual(bm_gap_sheet["B3"].number_format, "0.000")

            acc_sheet = workbook["ACC"]
            self.assertEqual(acc_sheet["A3"].value, "label_flipping_untargeted")
            self.assertEqual([acc_sheet.cell(row=2, column=col).value for col in range(2, 5)], ["原始", "提取+投影", "Base"])
            self.assertAlmostEqual(float(acc_sheet["B3"].value), 0.4567, places=4)
            self.assertAlmostEqual(float(acc_sheet["C3"].value), 0.5678, places=4)
            self.assertAlmostEqual(float(acc_sheet["D3"].value), 0.8123, places=4)
            self.assertEqual(acc_sheet["B3"].number_format, "0.000%")

            asr_sheet = workbook["ASR"]
            self.assertEqual(asr_sheet["A3"].value, "badnets")
            self.assertAlmostEqual(float(asr_sheet["B3"].value), 0.1234, places=4)
            self.assertAlmostEqual(float(asr_sheet["C3"].value), 0.2345, places=4)
            self.assertAlmostEqual(float(asr_sheet["D3"].value), 0.6789, places=4)
            self.assertEqual(asr_sheet["B3"].number_format, "0.000%")

    def test_export_experiment_results_can_save_bm_gap_only(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            self._export_toy_results(temp_dir, excel_metric_sheets="bm_gap")

            workbook = load_workbook(os.path.join(temp_dir, "ToyExperiment.xlsx"), data_only=False)
            self.assertEqual(workbook.sheetnames, ["BMGap"])

    def test_export_experiment_results_can_save_acc_asr_only(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            self._export_toy_results(temp_dir, excel_metric_sheets="acc_asr")

            workbook = load_workbook(os.path.join(temp_dir, "ToyExperiment.xlsx"), data_only=False)
            self.assertEqual(workbook.sheetnames, ["ACC", "ASR"])


if __name__ == "__main__":
    unittest.main()

from __future__ import annotations

import math
import os
import re
from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


def _slugify_tag(value: object) -> str:
    text = str(value).strip().lower()
    if not text:
        return "na"
    text = text.replace(".", "p")
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "na"


def _build_run_suffix(config) -> str:
    model = _slugify_tag(getattr(config, "MODEL", "model"))
    dataset = _slugify_tag(getattr(config, "DATASET", "dataset"))
    partition = getattr(config, "PARTITION", "partition")
    if isinstance(partition, dict):
        partition = partition.get("type", "partition")
    partition = _slugify_tag(partition)
    lr = _slugify_tag(getattr(config, "LR", "lr"))
    # 顶层结果目录不再把投毒比例编码进名称；
    # 不同 poison_rate 的结果统一下沉到子目录里分开存放。
    return f"{model}_{dataset}_{partition}_lr{lr}"


def _build_export_stem(
    experiment_name: str,
    poison_rate: object | None = None,
) -> str:
    # poison_rate 已经通过子目录隔离，这里不再重复写进文件名。
    _ = poison_rate
    return experiment_name


def create_run_directory(results_root: str, config=None) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if config is None:
        directory_name = timestamp
    else:
        directory_name = f"{timestamp}_{_build_run_suffix(config)}"
    output_dir = os.path.join(results_root, directory_name)
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def create_poison_rate_directory(parent_dir: str, poison_rate: object) -> str:
    # 不同投毒比例放到独立子目录，避免列表 sweep 时混在同一层。
    poison_rate_slug = _slugify_tag(poison_rate)
    output_dir = os.path.join(parent_dir, f"poison_rate_{poison_rate_slug}")
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


_REPORT_METHOD_ORDER = ["krum", "multi_krum", "median", "clustering"]
_BMGAP_FEATURE_LAYOUTS = [
    ("原始", ["raw_full"]),
    ("提取", ["selected_layers_balanced", "selected_layers"]),
    (
        "提取+投影",
        [
            "selected_layers_balanced_projected",
            "selected_layers_projected",
        ],
    ),
]
_PERFORMANCE_FEATURE_LAYOUTS = [
    ("原始", ["raw_full"]),
    (
        "提取+投影",
        [
            "selected_layers_balanced_projected",
            "selected_layers_projected",
        ],
    ),
]


def _resolve_report_methods(
    summary_records: List[Dict[str, object]],
    methods: List[str],
) -> List[str]:
    available_methods = {
        str(method).strip().lower()
        for method in methods
    }
    available_methods.update(
        str(record.get("method", "")).strip().lower()
        for record in summary_records
    )
    available_methods.discard("")
    available_methods.discard("fedavg")

    ordered_methods = [
        method_name
        for method_name in _REPORT_METHOD_ORDER
        if method_name in available_methods
    ]
    ordered_methods.extend(
        sorted(method_name for method_name in available_methods if method_name not in ordered_methods)
    )
    return ordered_methods


def _resolve_feature_layouts(
    summary_records: List[Dict[str, object]],
    feature_modes: List[str],
    layout_specs: List[tuple[str, List[str]]],
) -> List[tuple[str, str | None]]:
    available_feature_modes = {
        str(feature_mode).strip().lower()
        for feature_mode in feature_modes
    }
    available_feature_modes.update(
        str(record.get("feature_mode", "")).strip().lower()
        for record in summary_records
    )

    resolved_layouts: List[tuple[str, str | None]] = []
    for label, candidates in layout_specs:
        selected_feature_mode = next(
            (
                candidate
                for candidate in candidates
                if str(candidate).strip().lower() in available_feature_modes
            ),
            None,
        )
        resolved_layouts.append((label, selected_feature_mode))
    return resolved_layouts


def _filter_attacks_by_mode(
    summary_records: List[Dict[str, object]],
    attacks: List[str],
    attack_mode: str,
) -> List[str]:
    mode_by_attack: Dict[str, str] = {}
    for record in summary_records:
        attack_name = str(record.get("attack_name", ""))
        if not attack_name or attack_name in mode_by_attack:
            continue
        mode_by_attack[attack_name] = str(record.get("attack_mode", "")).strip().lower()

    normalized_mode = str(attack_mode).strip().lower()
    return [
        attack_name
        for attack_name in attacks
        if mode_by_attack.get(str(attack_name), "") == normalized_mode
    ]


def _build_metric_lookup(
    summary_records: List[Dict[str, object]],
    metric_key: str,
    fallback_metric_key: str | None = None,
) -> Dict[tuple[str, str, str], object]:
    lookup: Dict[tuple[str, str, str], object] = {}
    for record in summary_records:
        key = (
            str(record.get("attack_name", "")),
            str(record.get("method", "")).strip().lower(),
            str(record.get("feature_mode", "")).strip().lower(),
        )
        # 新增指标字段时保留回退键，避免旧 summary 记录导出 Excel 时整列空掉。
        lookup[key] = record.get(
            metric_key,
            record.get(fallback_metric_key, float("nan")) if fallback_metric_key is not None else float("nan"),
        )
    return lookup


def _resolve_excel_metric_sheets(excel_metric_sheets: object) -> List[str]:
    raw_value = "all" if excel_metric_sheets is None else excel_metric_sheets
    raw_items = [raw_value] if isinstance(raw_value, str) else list(raw_value)

    selected_sheets: List[str] = []
    for raw_item in raw_items:
        normalized_item = str(raw_item).strip().lower().replace("-", "_").replace(" ", "_")
        if normalized_item in {"all", "both"}:
            candidate_sheets = ["BMGap", "ACC", "ASR"]
        elif normalized_item in {"bm_gap", "bmgap", "bm"}:
            candidate_sheets = ["BMGap"]
        elif normalized_item in {"acc_asr", "acc_and_asr", "acc+asr", "acc/asr", "performance"}:
            candidate_sheets = ["ACC", "ASR"]
        elif normalized_item == "acc":
            candidate_sheets = ["ACC"]
        elif normalized_item == "asr":
            candidate_sheets = ["ASR"]
        else:
            raise ValueError(
                "Unsupported EXCEL_METRIC_SHEETS value: "
                f"{raw_item!r}. Expected 'all', 'bm_gap', or 'acc_asr'."
            )

        for sheet_name in candidate_sheets:
            # 保持用户给定顺序，同时去重，避免重复创建同名 sheet。
            if sheet_name not in selected_sheets:
                selected_sheets.append(sheet_name)

    if not selected_sheets:
        raise ValueError("EXCEL_METRIC_SHEETS 不能为空。")
    return selected_sheets


def _coerce_excel_number(value: object) -> float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _lookup_metric_value(
    metric_lookup: Dict[tuple[str, str, str], object],
    attack_name: str,
    method_name: str,
    feature_mode: str | None,
) -> float | None:
    normalized_method = str(method_name).strip().lower()
    if feature_mode is not None:
        value = metric_lookup.get(
            (
                str(attack_name),
                normalized_method,
                str(feature_mode).strip().lower(),
            ),
            float("nan"),
        )
        return _coerce_excel_number(value)

    # FedAvg 与特征模式无关；如果 raw_full 以外的模式被记录，也回退到同一攻击下的首个 FedAvg 结果。
    for (lookup_attack, lookup_method, _lookup_feature), value in metric_lookup.items():
        if lookup_attack == str(attack_name) and lookup_method == normalized_method:
            return _coerce_excel_number(value)
    return None


def _style_metric_sheet(worksheet, max_row: int, max_column: int, number_format: str) -> None:
    thin_side = Side(style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if cell.row <= 2 or cell.column == 1:
                cell.font = Font(bold=True)
            if cell.row >= 3 and cell.column >= 2:
                cell.number_format = number_format

    worksheet.column_dimensions["A"].width = 32
    for column_index in range(2, max_column + 1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = 15


def _write_metric_sheet(
    workbook: Workbook,
    sheet_name: str,
    attacks: List[str],
    methods: List[str],
    feature_layouts: List[tuple[str, str | None]],
    metric_lookup: Dict[tuple[str, str, str], object],
    method_display_names: Dict[str, str],
    number_format: str,
    include_fedavg_base: bool,
) -> None:
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.cell(row=2, column=1, value="attack")

    current_column = 2
    for method_name in methods:
        method_label = method_display_names.get(method_name, str(method_name))
        start_column = current_column
        for feature_label, _feature_mode in feature_layouts:
            worksheet.cell(row=2, column=current_column, value=feature_label)
            current_column += 1
        end_column = current_column - 1
        if end_column > start_column:
            worksheet.merge_cells(
                start_row=1,
                start_column=start_column,
                end_row=1,
                end_column=end_column,
            )
        worksheet.cell(row=1, column=start_column, value=method_label)

    if include_fedavg_base:
        fedavg_label = method_display_names.get("fedavg", "FedAvg")
        worksheet.cell(row=1, column=current_column, value=fedavg_label)
        worksheet.cell(row=2, column=current_column, value="Base")
        current_column += 1
    max_column = current_column - 1

    for row_offset, attack_name in enumerate(attacks, start=3):
        worksheet.cell(row=row_offset, column=1, value=attack_name)
        current_column = 2
        for method_name in methods:
            for _feature_label, feature_mode in feature_layouts:
                worksheet.cell(
                    row=row_offset,
                    column=current_column,
                    value=_lookup_metric_value(
                        metric_lookup=metric_lookup,
                        attack_name=attack_name,
                        method_name=method_name,
                        feature_mode=feature_mode,
                    ),
                )
                current_column += 1
        if include_fedavg_base:
            worksheet.cell(
                row=row_offset,
                column=current_column,
                value=_lookup_metric_value(
                    metric_lookup=metric_lookup,
                    attack_name=attack_name,
                    method_name="fedavg",
                    feature_mode=None,
                ),
            )

    max_row = max(2, len(attacks) + 2)
    max_column = max(1, max_column)
    _style_metric_sheet(
        worksheet=worksheet,
        max_row=max_row,
        max_column=max_column,
        number_format=number_format,
    )


def export_experiment_results(
    output_dir: str,
    experiment_name: str,
    summary_records: List[Dict[str, object]],
    round_logs: List[Dict[str, object]],
    krum_score_logs: List[Dict[str, object]],
    attacks: List[str],
    methods: List[str],
    feature_modes: List[str],
    method_display_names: Dict[str, str],
    feature_display_names: Dict[str, str],
    poison_rate: float | None = None,
    save_csv_logs: bool = True,
    save_round_logs: bool = True,
    export_excel: bool = True,
    excel_metric_sheets: object = "all",
) -> None:
    # 这些参数暂时保留在接口里，避免影响现有调用方；但当前版本不再导出对应文件。
    _ = round_logs, krum_score_logs, save_csv_logs, save_round_logs
    _ = feature_display_names
    report_methods = _resolve_report_methods(summary_records=summary_records, methods=methods)
    bm_gap_feature_layouts = _resolve_feature_layouts(
        summary_records=summary_records,
        feature_modes=feature_modes,
        layout_specs=_BMGAP_FEATURE_LAYOUTS,
    )
    performance_feature_layouts = _resolve_feature_layouts(
        summary_records=summary_records,
        feature_modes=feature_modes,
        layout_specs=_PERFORMANCE_FEATURE_LAYOUTS,
    )
    targeted_attacks = _filter_attacks_by_mode(
        summary_records=summary_records,
        attacks=attacks,
        attack_mode="targeted",
    )
    untargeted_attacks = _filter_attacks_by_mode(
        summary_records=summary_records,
        attacks=attacks,
        attack_mode="untargeted",
    )

    export_stem = _build_export_stem(
        experiment_name=experiment_name,
        poison_rate=poison_rate,
    )
    excel_path = os.path.join(output_dir, f"{export_stem}.xlsx")

    if export_excel:
        selected_metric_sheets = _resolve_excel_metric_sheets(excel_metric_sheets)
        workbook = Workbook()
        workbook.remove(workbook.active)
        for metric_sheet in selected_metric_sheets:
            # 按配置只创建需要的指标页，避免无关 sheet 干扰后续汇总。
            if metric_sheet == "BMGap":
                _write_metric_sheet(
                    workbook=workbook,
                    sheet_name="BMGap",
                    attacks=attacks,
                    methods=report_methods,
                    feature_layouts=bm_gap_feature_layouts,
                    metric_lookup=_build_metric_lookup(summary_records, metric_key="mean_bm_gap"),
                    method_display_names=method_display_names,
                    number_format="0.000",
                    include_fedavg_base=False,
                )
            elif metric_sheet == "ACC":
                _write_metric_sheet(
                    workbook=workbook,
                    sheet_name="ACC",
                    attacks=untargeted_attacks,
                    methods=report_methods,
                    feature_layouts=performance_feature_layouts,
                    metric_lookup=_build_metric_lookup(summary_records, metric_key="final_acc"),
                    method_display_names=method_display_names,
                    number_format="0.000%",
                    include_fedavg_base=True,
                )
            elif metric_sheet == "ASR":
                _write_metric_sheet(
                    workbook=workbook,
                    sheet_name="ASR",
                    attacks=targeted_attacks,
                    methods=report_methods,
                    feature_layouts=performance_feature_layouts,
                    metric_lookup=_build_metric_lookup(
                        summary_records,
                        metric_key="tail_mean_asr",
                        fallback_metric_key="final_asr",
                    ),
                    method_display_names=method_display_names,
                    number_format="0.000%",
                    include_fedavg_base=True,
                )
        workbook.save(excel_path)
